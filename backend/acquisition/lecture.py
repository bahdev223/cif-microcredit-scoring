"""Lecture des fichiers déposés par l'institution.

Un export de système d'information n'est jamais propre : encodage variable,
point-virgule plutôt que virgule, colonnes vides en fin de ligne, en-tête
précédé de lignes de titre. Ce module ramène tout cela à une même forme —
des en-têtes et des lignes — sans rien interpréter.

CSV et Excel sont traités ici. Aucune dépendance à pandas : openpyxl suffit,
et l'application doit pouvoir tourner sans la pile data science.
"""

import csv
from io import BytesIO, StringIO

ENCODAGES = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
SEPARATEURS = (";", ",", "\t", "|")
EXTENSIONS_TABLEUR = (".xlsx", ".xlsm")
LIGNES_ECHANTILLON = 5


class LectureImpossible(ValueError):
    """Le fichier n'a pas pu être lu, ou ne contient pas de tableau."""


def lire(fichier, feuille=None):
    """Retourne (en-têtes, lignes) pour un fichier CSV ou Excel.

    Les lignes sont des dictionnaires en-tête → valeur texte. La conversion des
    types n'est pas faite ici : elle appartient à la correspondance, qui seule
    sait quel champ attend une date ou un montant.
    """
    nom = (getattr(fichier, "name", "") or "").lower()
    contenu = fichier.read()
    if isinstance(contenu, str):
        contenu = contenu.encode("utf-8")
    if nom.endswith(EXTENSIONS_TABLEUR):
        return _lire_tableur(contenu, nom, feuille)
    return _lire_csv(contenu, nom)


def _decoder(contenu, nom):
    for encodage in ENCODAGES:
        try:
            return contenu.decode(encodage)
        except UnicodeDecodeError:
            continue
    raise LectureImpossible(f"{nom} : encodage non reconnu.")


def _detecter_separateur(premiere_ligne):
    """Retient le séparateur qui découpe la ligne d'en-tête en le plus de colonnes."""
    return max(SEPARATEURS, key=premiere_ligne.count)


def _lire_csv(contenu, nom):
    texte = _decoder(contenu, nom)
    lignes_brutes = [ligne for ligne in texte.splitlines() if ligne.strip()]
    if not lignes_brutes:
        raise LectureImpossible(f"{nom} est vide.")

    separateur = _detecter_separateur(lignes_brutes[0])
    lecteur = csv.reader(StringIO("\n".join(lignes_brutes)), delimiter=separateur)
    tableau = [ligne for ligne in lecteur if any(cellule.strip() for cellule in ligne)]
    if not tableau:
        raise LectureImpossible(f"{nom} ne contient aucune ligne exploitable.")
    return _structurer(tableau, nom)


def feuilles_tableur(fichier):
    """Retourne les feuilles non vides d'un fichier Excel, sans lire son tableau.

    L'agent peut ainsi choisir explicitement la feuille à analyser. Le premier
    onglet n'est jamais supposé être le bon : il peut être une page de garde.
    """
    nom = (getattr(fichier, "name", "") or "").lower()
    contenu = fichier.read()
    if not nom.endswith(EXTENSIONS_TABLEUR):
        return []
    try:
        from openpyxl import load_workbook
        classeur = load_workbook(BytesIO(contenu), read_only=True, data_only=True)
    except Exception as erreur:
        raise LectureImpossible(f"{nom} : fichier Excel illisible.") from erreur
    try:
        return [nom_feuille for nom_feuille in classeur.sheetnames
                if any(any(cellule is not None and str(cellule).strip() for cellule in ligne)
                       for ligne in classeur[nom_feuille].iter_rows(values_only=True))]
    finally:
        classeur.close()


def _lire_tableur(contenu, nom, feuille=None):
    try:
        from openpyxl import load_workbook
    except ImportError as erreur:  # pragma: no cover - dépendance déclarée
        raise LectureImpossible("La lecture des fichiers Excel nécessite openpyxl.") from erreur

    try:
        classeur = load_workbook(BytesIO(contenu), read_only=True, data_only=True)
    except Exception as erreur:
        raise LectureImpossible(f"{nom} : fichier Excel illisible.") from erreur
    nom_feuille = feuille or classeur.sheetnames[0]
    if nom_feuille not in classeur.sheetnames:
        classeur.close()
        raise LectureImpossible(f"La feuille « {nom_feuille} » est introuvable dans {nom}.")
    feuille_excel = classeur[nom_feuille]
    tableau = []
    for ligne in feuille_excel.iter_rows(values_only=True):
        cellules = ["" if cellule is None else str(cellule).strip() for cellule in ligne]
        if any(cellules):
            tableau.append(cellules)
    classeur.close()
    if not tableau:
        raise LectureImpossible(f"{nom} ne contient aucune ligne exploitable.")
    return _structurer(tableau, nom)


def _structurer(tableau, nom):
    """Identifie la ligne d'en-tête et construit les lignes de données.

    Certains exports font précéder le tableau d'un titre ou d'une date. On
    retient comme en-tête la première ligne qui porte au moins deux libellés
    non vides et non numériques.
    """
    index_entete = 0
    for index, ligne in enumerate(tableau[:10]):
        remplies = [cellule.strip() for cellule in ligne if cellule.strip()]
        if len(remplies) >= 2 and not all(_est_nombre(cellule) for cellule in remplies):
            index_entete = index
            break

    entetes_brutes = tableau[index_entete]
    entetes, vues = [], {}
    for position, entete in enumerate(entetes_brutes):
        libelle = entete.strip() or f"colonne_{position + 1}"
        if libelle in vues:
            vues[libelle] += 1
            libelle = f"{libelle}_{vues[libelle]}"
        else:
            vues[libelle] = 1
        entetes.append(libelle)

    lignes = []
    for ligne in tableau[index_entete + 1:]:
        valeurs = list(ligne) + [""] * (len(entetes) - len(ligne))
        lignes.append({entete: str(valeurs[position]).strip() for position, entete in enumerate(entetes)})

    if not lignes:
        raise LectureImpossible(f"{nom} ne contient pas de ligne de données sous l'en-tête.")
    return entetes, lignes


def _est_nombre(valeur):
    try:
        float(valeur.replace(" ", "").replace(",", "."))
        return True
    except (ValueError, AttributeError):
        return False


def echantillon(lignes, entetes, taille=LIGNES_ECHANTILLON):
    """Quelques lignes réelles, pour que l'agent voie ce qu'il associe."""
    return [{entete: ligne.get(entete, "") for entete in entetes} for ligne in lignes[:taille]]
